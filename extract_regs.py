#!/usr/bin/env python3
"""Extract register descriptions from CANFD translation and add to Excel."""

import re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

md_path = "/Users/ai-work/ai-dv/agents/claw-pdf/workspace/canfd/pg223_canfd_中文翻译.md"
xlsx_path = "/Users/ai-work/ai-dv/agents/claw-pdf/workspace/canfd/CANFD_IP_验证功能点.xlsx"

with open(md_path, 'r') as f:
    content = f.read()

start = content.find("## 内核寄存器描述")
end = content.find("\n## ", start + 100)
if end == -1:
    end = len(content)
reg_section = content[start:end]

entries = []
current_title = ""
current_lines = []

for line in reg_section.split('\n'):
    if line.startswith('### ') and '寄存器' in line:
        if current_title and current_lines:
            entries.append((current_title, '\n'.join(current_lines)))
        current_title = line.replace('### ', '').strip()
        current_lines = []
    elif current_title:
        current_lines.append(line)

if current_title and current_lines:
    entries.append((current_title, '\n'.join(current_lines)))

def parse_register_entry(title, body):
    result = {'title': title, 'addr': '', 'name_en': '', 'name_cn': '',
              'description': '', 'table_bits': [], 'notes': []}
    
    addr_match = re.search(r'地址偏移\s*\+\s*(0x[0-9A-Fa-f]+)', title)
    if addr_match:
        result['addr'] = addr_match.group(1)
    
    name_match = re.search(r'（([A-Za-z0-9_ ,/+\-]+?)，?\s*(地址偏移|邮箱|RX|TX)', title)
    if name_match:
        result['name_en'] = name_match.group(1).strip()
    else:
        name_simple = re.match(r'([^(（]+)', title)
        if name_simple:
            result['name_cn'] = name_simple.group(1).strip()
    
    in_table = False
    table_rows = []
    for line in body.split('\n'):
        line = line.strip()
        if line.startswith('|') and '---' not in line:
            if not in_table:
                in_table = True
            else:
                parts = [p.strip() for p in line.split('|')[1:-1]]
                if len(parts) >= 2:
                    cleaned = [re.sub(r'<[^>]+>', ' ', p).strip() for p in parts]
                    table_rows.append(cleaned)
        else:
            in_table = False
    
    result['table_rows'] = table_rows
    
    pre_table = body
    if '|' in body:
        first_table = body.index('|')
        pre_table = body[:first_table]
    desc_lines = [l.strip() for l in pre_table.split('\n') if l.strip() and not l.startswith('#') and not l.startswith('**') and not l.startswith('>')]
    result['description'] = ' '.join(desc_lines[:5])
    
    notes = []
    for line in body.split('\n'):
        ls = line.strip()
        if ls.startswith('>'):
            notes.append(re.sub(r'^>\s*', '', ls))
    result['notes'] = notes
    
    return result

parsed = [parse_register_entry(title, body) for title, body in entries]

# ===== Create Excel =====
wb = openpyxl.load_workbook(xlsx_path)
if '寄存器详细描述' in wb.sheetnames:
    del wb['寄存器详细描述']

ws = wb.create_sheet('寄存器详细描述', 0)

header_font = Font(name="Arial Unicode MS", bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
reg_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
normal_font = Font(name="Arial Unicode MS", size=9)
bold_font = Font(name="Arial Unicode MS", size=9, bold=True)
small_font = Font(name="Arial Unicode MS", size=8)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

ws.merge_cells('A1:F1')
title_cell = ws.cell(row=1, column=1, value="CANFD Controller IP 寄存器详细描述 (PG223 v3.0 中文翻译)")
title_cell.font = Font(name="Arial Unicode MS", bold=True, size=13, color="1F3864")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

col_headers = ["地址", "寄存器缩写", "寄存器中文名", "字段说明", "描述", "注意事项"]
col_widths = [12, 22, 30, 60, 45, 35]

for col, (h, w) in enumerate(zip(col_headers, col_widths), 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A3"

row = 3
count = 0
for i, reg in enumerate(parsed):
    if not reg['addr']:
        continue
    count += 1
    
    field_parts = []
    for tbl_row in reg.get('table_rows', []):
        if len(tbl_row) >= 3:
            bit_pos = tbl_row[0] if len(tbl_row) > 0 else ''
            bit_name = tbl_row[1] if len(tbl_row) > 1 else ''
            bit_desc = tbl_row[-1] if len(tbl_row) > 2 else ''
            bit_desc = re.sub(r'<[^>]+>', '', bit_desc).replace('\n', ' ').strip()
            if bit_pos:
                field_parts.append(f"[{bit_pos}] {bit_name}: {bit_desc[:100]}")
    
    field_text = '\n'.join(field_parts[:12]) if field_parts else reg['description'][:200]
    note_text = '\n'.join(reg.get('notes', [])[:4]) if reg.get('notes', False) else ''
    name_en = reg.get('name_en', '')
    name_cn = reg.get('name_cn', '')
    
    if not name_en:
        title_parts = reg['title'].split('（', 1)
        if len(title_parts) >= 2:
            name_cn = title_parts[0].strip()
            inner = title_parts[1].split('）')[0].split(')')[0]
            en_match = re.findall(r'[A-Z][A-Z_0-9a-z]+', inner)
            name_en = ', '.join(en_match[:2]) if en_match else inner[:40]
        else:
            name_cn = reg['title'].split('（')[0].split('(')[0].strip()
    
    desc = reg['description'][:300]
    
    data = [reg['addr'], name_en, name_cn, field_text, desc, note_text]
    
    bg = reg_fill if i % 2 == 0 else None
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = small_font if col >= 4 else bold_font
        cell.alignment = wrap_align if col >= 4 else center_align
        cell.border = thin_border
        if bg:
            cell.fill = bg
    
    max_lines = max(field_text.count('\n'), desc.count('\n'), note_text.count('\n'), 1)
    ws.row_dimensions[row].height = max(25, min(250, max_lines * 15 + 10))
    row += 1

wb.save(xlsx_path)
print(f"OK: Added '寄存器详细描述' sheet with {count} registers")
